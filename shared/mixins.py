import io
import json

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.views import redirect_to_login
from django.core.exceptions import PermissionDenied
from django.db.models import ProtectedError
from django.http import HttpResponse
from django.shortcuts import redirect


class PermissionMixin(LoginRequiredMixin, PermissionRequiredMixin):
    """
    Redirige a login si el usuario no está autenticado.
    Devuelve 403 si está autenticado pero le falta el permiso requerido.

    Uso en CBVs:
        class MiVista(PermissionMixin, ListView):
            permission_required = 'app.view_model'
    """

    def handle_no_permission(self):
        if self.request.user.is_authenticated:
            raise PermissionDenied
        return redirect_to_login(
            self.request.get_full_path(),
            self.get_login_url(),
            self.get_redirect_field_name(),
        )


class ProtectedDeleteMixin:
    """Evita que un DeleteView tumbe la página con un error 500 cuando el
    registro no se puede borrar porque otro (factura, compra, etc.) lo
    referencia con on_delete=PROTECT. Muestra un mensaje claro en su lugar.

    Uso: agregarlo primero en el MRO, junto a PermissionMixin y DeleteView:
        class ProductDeleteView(ProtectedDeleteMixin, PermissionMixin, DeleteView):
            ...
    """

    def form_valid(self, form):
        try:
            return super().form_valid(form)
        except ProtectedError:
            messages.error(
                self.request,
                f'No se puede eliminar "{self.object}" porque está siendo utilizado en otros registros del sistema '
                '(por ejemplo, facturas o compras).'
            )
            return redirect(self.get_success_url())


class StaffRequiredMixin(LoginRequiredMixin):
    """
    Mixin que verifica si el usuario es miembro del staff.
    Si no es staff, redirige con mensaje de error.

    Uso:
        class BrandDeleteView(LoginRequiredMixin, StaffRequiredMixin, DeleteView):
            ...

    ¿Por qué?
        Porque solo el personal autorizado (staff) debe poder
        eliminar registros. Un usuario normal puede ver y crear,
        pero no borrar información importante del sistema.

    ¿Cómo funciona?
        1. El usuario intenta acceder a una vista protegida.
        2. dispatch() se ejecuta ANTES que la vista.
        3. Si user.is_staff es False → redirige con mensaje de error.
        4. Si user.is_staff es True → ejecuta la vista normalmente.
    """

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        if not request.user.is_staff:
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('billing:home')
        return super().dispatch(request, *args, **kwargs)


def _resolve_field(obj, field):
    if callable(field):
        return field(obj)
    value = obj
    for part in field.split('.'):
        value = getattr(value, part, '')
        if callable(value):
            value = value()
    return '' if value is None else value


def _build_export_rows(queryset, fields):
    headers = [label for label, _ in fields]
    rows = [[str(_resolve_field(obj, f)) for _, f in fields] for obj in queryset]
    return headers, rows


def _export_pdf_response(title, headers, rows, col_widths=None):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    styles = getSampleStyleSheet()
    hdr_style = ParagraphStyle(
        'PDFHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        textColor=colors.white,
        leading=10,
    )
    cell_style = ParagraphStyle(
        'PDFCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=7,
        leading=9,
        wordWrap='LTR',
    )

    # Wrap every cell in Paragraph so ReportLab breaks lines automatically
    header_row = [Paragraph(h, hdr_style) for h in headers]
    data_rows  = [
        [Paragraph(str(cell), cell_style) for cell in row]
        for row in rows
    ]

    # Column widths: use col_widths (relative weights) if defined
    page_w = landscape(A4)[0]
    margin = 1 * cm
    available_w = page_w - 2 * margin
    n_cols = max(len(headers), 1)
    if col_widths and len(col_widths) == n_cols:
        total = sum(col_widths)
        col_widths_px = [available_w * w / total for w in col_widths]
    else:
        col_widths_px = [available_w / n_cols] * n_cols

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=margin, rightMargin=margin,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )

    table = Table([header_row] + data_rows, colWidths=col_widths_px, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1,  0), colors.HexColor('#343a40')),
        ('TEXTCOLOR',     (0, 0), (-1,  0), colors.white),
        ('ALIGN',         (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN',        (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#dee2e6')),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING',   (0, 0), (-1, -1), 5),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 5),
    ]))

    doc.build([
        Paragraph(title, styles['Title']),
        Spacer(1, 0.3 * cm),
        table,
    ])
    buf.seek(0)
    fname = title.replace(' ', '_').lower() + '.pdf'
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


def _export_excel_response(title, headers, rows):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    # ── Estilos ──────────────────────────────────────────────
    thin      = Side(border_style='thin', color='DEE2E6')
    brd       = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill  = PatternFill(fill_type='solid', fgColor='F0F4FF')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    dat_align = Alignment(horizontal='left',   vertical='top',    wrap_text=True)

    n_cols = len(headers)

    # ── Fila 1: título del listado ────────────────────────────
    title_cell = ws.cell(row=1, column=1, value=title.upper())
    title_cell.font      = Font(bold=True, size=13, color='FFFFFF')
    title_cell.fill      = PatternFill(fill_type='solid', fgColor='0D6EFD')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32
    if n_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    # ── Fila 2: encabezados de columna ────────────────────────
    ws.row_dimensions[2].height = 28
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font      = Font(bold=True, color='FFFFFF', size=10)
        cell.fill      = PatternFill(fill_type='solid', fgColor='343A40')
        cell.alignment = hdr_align
        cell.border    = brd

    # ── Filas de datos (desde fila 3) ────────────────────────
    for idx, row in enumerate(rows):
        r_idx = idx + 3
        ws.row_dimensions[r_idx].height = 42
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=str(val))
            cell.border    = brd
            cell.alignment = dat_align
            if idx % 2 == 1:
                cell.fill = alt_fill

    # ── Anchos de columna (min 12, max 52) ───────────────────
    for c_idx in range(1, n_cols + 1):
        col_letter = get_column_letter(c_idx)
        col_vals   = [ws.cell(row=r, column=c_idx).value or ''
                      for r in range(2, ws.max_row + 1)]
        best       = max(len(str(v)) for v in col_vals)
        ws.column_dimensions[col_letter].width = min(max(best + 3, 12), 52)

    # ── Congelar título + encabezado ─────────────────────────
    ws.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = title.replace(' ', '_').lower() + '.xlsx'
    response = HttpResponse(
        buf,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


def export_queryset(request, queryset, title, fields, col_widths=None):
    """Exporta un queryset a PDF/Excel según ?export=pdf|excel&col=<etiqueta>.

    Pensado para reutilizarse tanto desde `ExportMixin` (CBVs) como desde
    vistas basadas en función. Devuelve None si la request no pide exportar,
    para que el llamador continúe con el flujo normal (render de la lista).

    fields: [(etiqueta_columna, ruta_attr_o_callable), ...] — la ruta admite
    notación de punto ('brand.name') y los callables reciben el objeto fila.
    col_widths: pesos relativos de columna (mismo orden y longitud que fields),
    solo usados en la exportación a PDF.
    """
    fmt = request.GET.get('export', '').lower()
    if fmt not in ('pdf', 'excel'):
        return None
    if not request.user.has_perm('billing.can_export'):
        messages.error(request, 'No tienes permiso para exportar reportes a PDF/Excel.')
        return None
    cols = request.GET.getlist('col')
    active = [(lbl, attr) for lbl, attr in fields if lbl in cols] if cols else fields
    active = active or fields
    headers, rows = _build_export_rows(queryset, active)
    if fmt == 'pdf':
        weights = None
        if col_widths and len(col_widths) == len(fields):
            all_labels = [lbl for lbl, _ in fields]
            weights = [col_widths[all_labels.index(lbl)] for lbl, _ in active if lbl in all_labels]
        return _export_pdf_response(title, headers, rows, weights)
    return _export_excel_response(title, headers, rows)


class ExportMixin:
    """Agrega exportación PDF/Excel a cualquier ListView via ?export=pdf o ?export=excel.

    Atributos que debe definir la subclase:
        export_title  (str)  — título del documento
        export_fields (list) — [(etiqueta_columna, ruta_attr_o_callable), ...]
        export_col_widths (list, opcional) — pesos relativos de columna para el PDF

    La ruta de atributo soporta notación de punto ('brand.name').
    Los callables reciben el objeto fila y devuelven el valor a mostrar.
    Usa self.object_list (queryset completo, sin paginar) para la exportación.
    """

    export_title = 'Listado'
    export_fields = []  # [(label, attr_path_or_callable), ...]
    export_col_widths = None

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['export_field_labels_json'] = json.dumps(
            [lbl for lbl, _ in self.export_fields], ensure_ascii=False
        )
        return ctx

    def render_to_response(self, context, **response_kwargs):
        response = export_queryset(
            self.request,
            getattr(self, 'object_list', self.get_queryset()),
            self.export_title,
            self.export_fields,
            self.export_col_widths,
        )
        if response is not None:
            return response
        return super().render_to_response(context, **response_kwargs)
